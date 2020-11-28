from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy
from difflib import SequenceMatcher

def str_similarity(a, b):
    return SequenceMatcher(None, a, b).ratio()

def wrap_text(dir):
    '''
    Scales the cell widths to improve readability.
    '''

    workbook = load_workbook(dir)
    worksheet = workbook.active

    dimensions = {}
    for row in worksheet.rows:
        for cell in row:
            if cell.value:
                # Calculate the widths for each column
                dimensions[cell.column_letter] = max(
                    (dimensions.get(cell.column_letter, 0),
                    min((
                        len(str(cell.value))),
                        20
                    ))
                )
            
            cell.alignment = Alignment(horizontal='left')

    # Apply the widths for each column
    for column, value in dimensions.items():
        worksheet.column_dimensions[column].width = value
    
    workbook.save(dir)

def add_headers(dir, types):
    '''
    Adds headers for 'Soort' and removes the 'Soort' column.
    '''

    workbook = load_workbook(dir)
    worksheet = workbook.active

    last_type = ''
    offset = 1
    for i in range(types.shape[0]):
        index = i + offset
        _type = types.iloc[i]
        
        # Prevent duplicate headers by using the str_similarity function
        if str_similarity(_type, last_type) < 0.9:
            # Add a header

            worksheet.move_range(
                'A${0}:D${1}'.format(index + 1, worksheet.max_row),
                rows=2
            )

            worksheet.merge_cells('A${0}:D${0}'.format(index + 2))

            cell = worksheet['A${0}'.format(index + 2)]
            cell.value = _type
            cell.font = copy(worksheet['A1'].font)

            thin = Side(border_style="thin", color="000000")
            cell.border = Border(
                top=thin,
                bottom=thin,
                left=thin,
                right=thin
            )

            last_type = _type
            offset += 2
    
    workbook.save(dir)