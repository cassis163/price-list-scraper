from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

def fit(dir):
    workbook = load_workbook(dir)
    worksheet = workbook.active

    dimensions = {}
    for row in worksheet.rows:
        for cell in row:
            if cell.value:
                dimensions[cell.column_letter] = max(
                    (dimensions.get(cell.column_letter, 0),
                    min((
                        len(str(cell.value))),
                        20
                    ))
                )
            
            align = 'left'

            cell.alignment = Alignment(horizontal=align)

    for column, value in dimensions.items():
        worksheet.column_dimensions[column].width = value
    
    workbook.save(dir) 